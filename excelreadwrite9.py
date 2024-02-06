from tkinter import filedialog
import tkinter as tk
import pandas as pd
import openpyxl, datetime
import sys
import collections

def main():
    root = tk.Tk()
    root.geometry("400x400")
    root.overrideredirect(True)

    title_bar = tk.Frame(root, bg="blue", height=30)
    title_bar.pack(fill=tk.X)

    label = tk.Label(title_bar, text="Laiko grafikų paruošimo aplikacija", fg="white", bg="blue")
    label.pack(side=tk.LEFT)

    close_button = tk.Button(title_bar, text="X", command=root.destroy, bg="red", fg="white")
    close_button.pack(side=tk.RIGHT)

    button = tk.Button(root, text="Perskaityti .xls grafiką", fg="white", bg="blue", command=browse_read_file)
    button.pack(ipadx=50, expand=True, anchor=tk.CENTER)
    button1 = tk.Button(root, text="Išeiti", fg="white", bg="blue", command=root.destroy)
    button1.pack(ipadx=92, expand=True, anchor=tk.N)

    title_bar.bind("<B1-Motion>", lambda event: move_window(event, root))
    title_bar.bind("<Button-1>", lambda event: on_title_bar_press(event, root))

    label.bind("<B1-Motion>", lambda event: move_window(event, root))
    label.bind("<Button-1>", lambda event: on_title_bar_press(event, root))

    root.mainloop()

    return "Success"

def move_window(event, root):
    root.geometry(f"+{root.winfo_pointerx() - initial_x}+{root.winfo_pointery() - initial_y}")

def on_title_bar_press(event, root):
    global initial_x, initial_y
    initial_x = event.x
    initial_y = event.y

def browse_read_file():
    textdata8, groups, workers_list_full, template_object_list, templates = [], [], [], [], []
    days = 0
    string = "end of code"
    data = {}
    filename = filedialog.askopenfilename(initialdir="/", title="Select a File", filetypes=[("Excel files", "*.xls")])
    
    if filename:
        textdata8 = read_datafile(filename)
        
    days = days_in_a_month()
    groups = find_groups(textdata8)
    workers_list_full = find_workers(groups, textdata8, days)
    template_object_list = find_templates_by_date_for_workers(days, workers_list_full)
    data, data1, data_templates = formatdataframe(days, workers_list_full, template_object_list)
    string = writetoexcel(data, data1, data_templates)

    return "Success"

# Nuskaitomi ir sutvarkomi laiko tvarkaraščio duomenys
def read_datafile(new_filename):
    start_index, end_index= 1, 1
    desiredline = ""
    textdata8, textdata3= [], []
    
    with open(new_filename, 'r', encoding='utf-8') as f:
        textdata = f.readlines()

    for line in textdata:
        if '/Data' in line:
            start_index = line.find('<Data ss:Type="String">') + len('<Data ss:Type="String">')
            end_index = line.find('<', start_index)
            desiredline = line[start_index:end_index]
            textdata8.append(desiredline.replace(">", "").replace("&#10;", " "))
            if desiredline != '':
                textdata3.append(desiredline)

    return textdata8

# Randami darbuotojai
def find_workers(groups, textdata8, days):
    k, a = 0, 0
    group = ""
    workers_list, current_group, sorted_unique_workers, worker_name_surname, only_worker_list, workers_list_full= [], groups[0], [], [], [], []
    time2 = get_time2()

    for i, line in enumerate(textdata8):
        if line != "":
            if line[0].isalpha():
                if i + 4 < len(textdata8):
                    if textdata8[i+1] != "" and textdata8[i-1] !="" and textdata8[i-2] !="":
                        if textdata8[i+1].isdigit() and textdata8[i-1].isdigit() and textdata8[i-2].isdigit():
                            k = i + 1 + 31 + 2 #pagal duota grafika visad yra 31 irasas, pabaigoje nuskusti du likusius days pakeiciama i 31
                            a = k
                            only_worker_list.append(line)

                            worker_name_surname.append(textdata8[i-1])
                            worker_name_surname.append(line)

                            # print(len(line.split()), "--------------------", line.split()[0], line.split()[1])

                            # if len(line.split()) == 3:
                            #     print(len(line.split()), "--------------------", line.split()[0] + " " + line.split()[1] + " " + line.split()[2])


                            if len(line.split()) == 3:
                                worker1 = Worker(tabel_number=textdata8[i-1], name=line.split()[0] , surname=line.split()[1] + " " + line.split()[2], time_start=[], time_end=[], time=[], division='')


                            if len(line.split()) == 2:
                                worker1 = Worker(tabel_number=textdata8[i-1], name=line.split()[0] , surname=line.split()[1], time_start=[], time_end=[], time=[], division='')
                            
                            #print("---", worker1.tabel_number, worker1.name, worker1.surname, "---")

                            try:
                                while (k < a + 31):
                                    if i + 31 < len(textdata8):
                                        k = k + 1
                                        worker1.time_start.append(textdata8[k])
                                        worker1.time_end.append(textdata8[k + 31])
                                        if textdata8[k] != '':
                                            if len(textdata8[k]) == 4:
                                                textdata8[k] = '0' + textdata8[k]
                                            if len(textdata8[k + 31]) == 4:
                                                textdata8[k + 31] = '0' + textdata8[k + 31]
                                            worker1.time.append(textdata8[k] + '-' + textdata8[k + 31])
                                        else:
                                            worker1.time.append('')

                                try:
                                    if days == 30:
                                        worker1.time.pop()
                                    if days == 29:
                                        worker1.time.pop()
                                        worker1.time.pop()
                                    if days == 28:
                                        worker1.time.pop()
                                        worker1.time.pop()
                                        worker1.time.pop()
                                except IndexError:
                                    print("laikai nepašalinti teisingai")
                                
                                workers_list.append(worker1)
                            except IndexError:
                                print("laikai neįrašyti teisingai")
                            
                            
                            
                    elif textdata8[i+4] != "" and textdata8[i-1] != "":
                        if textdata8[i+4][0].isalpha() and textdata8[i-1].isdigit():
                            worker_name_surname.append(line)
                    else:
                        print("String is not a groups or a worker")

    #Pritaikoma laikų suradimui Shift not found by title '07:30-15:00' in sheet 'Šablonai' row '2'
    sorted_unique_workers = sorted(list(set(workers_list)), key=lambda worker: int(worker.tabel_number))
    

    for i, line in enumerate(worker_name_surname):
        if line not in groups and line.isdigit() == False:
            tabel_nr = worker_name_surname[i-1]
            group = find_current_group(current_group)
            
            worker_full = Worker_Modified(tabel_number=tabel_nr, name=line.split(' ')[0], surname=find_worker_surname(sorted_unique_workers, tabel_nr), time_start=0, time_end=0, time= find_worker_time(sorted_unique_workers, tabel_nr),division=group, template = '', value = 'Pamaina/laikas')
            worker_full1 = Worker_Modified(tabel_number=tabel_nr, name=line.split(' ')[0] , surname=find_worker_surname(sorted_unique_workers, tabel_nr),  time_start=0, time_end=0, time= time2, division=find_current_group(current_group), template = '', value = 'Simbolis')
            workers_list_full.append(worker_full)
            workers_list_full.append(worker_full1)
        if line in groups and line != current_group:
            current_group = line

    for worker in workers_list_full:
        if worker.tabel_number == "1186":
            worker.tabel_number = "1044"

    workers_list_full.sort(key=lambda worker: int(worker.tabel_number))

    return workers_list_full


# Randamos šventines metų dienas kas antram worker objektui. Taip daroma siekiant pagal šabloną įrašyti reikšmes
def get_time2():
    current_date = datetime.datetime.now()
    current_month = current_date.month
    year_now, month_now = get_current_year_and_month()

    if current_month == 1:
        time2 = ['S', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
    if current_month == 2 and is_leap_year(year_now) == True:
        time2 = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'S', '', '', '', '', '', '', '', '', '', '', '', '', '']
    elif current_month == 2 and is_leap_year(year_now) == False:
        time2 = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'S', '', '', '', '', '', '', '', '', '', '', '', '']
    if current_month == 3:
        time2 = ['', '', '', '', '', '', '', '', '', '', 'S', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
    if current_month == 4:
        time2 = ['S', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
    if current_month == 5:
        time2 = ['S', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
    if current_month == 6:
        time2 = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'S', '', '', '', '', '', '']
    if current_month == 7:
        time2 = ['', '', '', '', '', 'S', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
    if current_month == 8:
        time2 = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', 'S', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
    if current_month == 9:
        time2 = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
    if current_month == 10:
        time2 = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
    if current_month == 11:
        time2 = ['S', 'S', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
    if current_month == 12:
        time2 = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'S', 'S', 'S', '', '', '', '', '']

    return time2


# Darbuotojų duomenys įrašomi į excel
def writetoexcel(data, data1, data_templates):
    output_file = "naujausias_tvarkarastis.xlsx" 
    source_file = "pamainos.xlsx"
    output_file1 = "vienas_asmuo.xlsx" # Kuriamas kintamasis vieno asmens įrašymui
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    # writer1 = pd.ExcelWriter(output_file1, engine='openpyxl') # Kuriamas kintamasis vieno asmens įrašymui
    workbook = writer.book
    # workbook1 = writer1.book # Kuriamas kintamasis vieno asmens įrašymui
    workbook.create_sheet('Pamainos')
    workbook.create_sheet('Šablonai')
    # workbook1.create_sheet('Pamainos') # Kuriamas kintamasis vieno asmens įrašymui
    # workbook1.create_sheet('Šablonai') # Kuriamas kintamasis vieno asmens įrašymui

    worksheet2 = writer.sheets['Pamainos']
    worksheet3 = writer.sheets['Šablonai']

    # worksheet2_1 = writer1.sheets['Pamainos'] # Kuriamas kintamasis vieno asmens įrašymui
    # worksheet3_2 = writer1.sheets['Šablonai'] # Kuriamas kintamasis vieno asmens įrašymui
    
    source_df1 = pd.read_excel(source_file, sheet_name='Pamainos')
    # source_df2 = pd.read_excel(source_file, sheet_name='Šablonai') 
    
    source_df1.to_excel(writer, sheet_name='Pamainos', index=False)

    
    
    new_font = openpyxl.styles.Font(name='Aptos Narrow', size=11)
    column_widths = {
            'A': 13, 'B': 11, 'C': 14, 'D': 30, 'E': 14, 'F': 14,
            'G': 13, 'H': 13, 'I': 13, 'J': 13, 'K': 13, 'L': 13,
            'M': 13, 'N': 13, 'O': 13, 'P': 13, 'Q': 13, 'R': 13,
            'S': 13, 'T': 13, 'U': 13, 'V': 13, 'W': 13, 'X': 13,
            'Y': 13, 'Z': 13, 'AA': 13, 'AB': 13, 'AC': 13, 'AD': 13,
            'AE': 13, 'AF': 13, 'AG': 13, 'AH': 13, 'AI': 13, 'AJ': 13,
            'AK': 13
        }
    column_widths2 = {
            'A': 15, 'B': 14, 'C': 14, 'D': 11, 'E': 14, 'F': 17
        }
    column_widths3 = {
            'A': 30, 'B': 2, 'C': 10, 'D': 10, 'E': 10, 'F': 10,
            'G': 2, 'H': 10, 'I': 10, 'J': 10, 'K': 10, 'L': 10,
            'M': 10, 'N': 2, 'O': 10, 'P': 10, 'Q': 10, 'R': 10,
            'S': 10, 'T': 10, 'U': 2, 'V': 10, 'W': 10, 'X': 10,
            'Y': 10, 'Z': 10, 'AA': 10, 'AB': 2, 'AC': 10, 'AD': 10,
            'AE': 10, 'AF': 10
        }


    for row in worksheet2.iter_rows():
        for index, cell in enumerate(row):
            if row[0].row == 1:
                cell.alignment = openpyxl.styles.Alignment(horizontal='left')
                cell.font = openpyxl.styles.Font(bold=False)
            cell.font = new_font
        for cell in row:
            cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin', color='FF000000'))

    for col, width in column_widths2.items():
        worksheet2.column_dimensions[col].width = width
    

    df2 = pd.DataFrame(data_templates)
    df2.to_excel(writer, sheet_name='Šablonai', index=False)
    worksheet_templates = writer.sheets['Šablonai']


    for row in worksheet_templates.iter_rows():
        for index, cell in enumerate(row):
            if row[0].row == 1:
                cell.alignment = openpyxl.styles.Alignment(horizontal='left')
                cell.font = openpyxl.styles.Font(bold=False)
            if index > 1:
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

            cell.font = new_font
            if row[0].row % 2 != 0:
                for cell in row:
                    cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='medium', color='FF000000'))
        
    for col, width in column_widths3.items():
        worksheet_templates.column_dimensions[col].width = width


    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='Tvarkaraštis', index=False)
    worksheet = writer.sheets['Tvarkaraštis']
    for row in worksheet.iter_rows():
            for index, cell in enumerate(row):
                if row[0].row == 1:
                    if index < 6:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='left')
                        cell.font = openpyxl.styles.Font(bold=False)
                    elif 6 <= index:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                        cell.font = openpyxl.styles.Font(bold=False)
                else:
                    if index == 0:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                    elif index > 0:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='left')
                cell.font = new_font

            if row[0].row % 2 != 0:
                for cell in row:
                    cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='medium', color='FF000000'))

    for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

    writer._save()

    return "Success"


# Randami šablonai. Laikai yra pasikartojantys, rasti kiek kartų pasikartoja laikai ir pagal tai atrinkti šablonus
# Į Work Sober sistemą internete suvedžiau naujus šablonus, nes be jų nepriima duomenų. Šablonai tokie:
# MK 2 pam. 6.00-14.30, 14.30-23.00
# 7.30-16.00
# 14.00-22.00, 6.00-14.00
# 6.00-14.00, 14.00-22.00
# 7.30-14.00, 10.00-18.30, 13.00-21.30
# 7.00-16.00, 10.00-18.30, 13.00-21.30
# Nepamišti naujų šablonų įtraukti į šablonų excelio puslapį
def find_templates_by_date_for_workers(days, workers_list_full):
    template_object_list_temp, template_object_list_temp_par = [], []
    x = 0

    times_list_with_count = ['07:00-16:00', 0, #1
                             '08:00-12:00', 0, #3
                             '08:00-16:30', 0, #5
                             '08:00-17:00', 0, #7
                             '08:00-15:45', 0, #9
                             '07:30-16:00', 0, #11
                             '06:00-14:00', 0, #13
                             '22:00-06:00', 0, #15
                             '14:00-22:00', 0, #17
                             '14:30-23:00', 0, #19
                             '10:00-18:30', 0, #21
                             '06:00-13:30', 0, #23
                             '06:00-14:30', 0, #25
                             '13:00-21:30', 0, #27
                             '08:00-16:00', 0, #29
                             '14:00-21:00', 0, #31
                             '07:00-15:00', 0, #33
                             '07:30-15:00', 0, #35
                             '10:00-17:30', 0, #37
                             '13:00-20:30', 0, #39
                             '14:30-22:00', 0] #41

    

    for i, worker in enumerate(workers_list_full):
        while days > x:
            # if len(worker.time) != 31:
            #     print(worker.name, worker.surname, worker.time)
            #     break
            if worker.time[x] == '07:00-16:00':
                times_list_with_count[1]+=1
            if worker.time[x] == '08:00-12:00':
                times_list_with_count[3]+=1
            if worker.time[x] == '08:00-16:30':
                times_list_with_count[5]+=1
            if worker.time[x] == '08:00-17:00':
                times_list_with_count[7]+=1
            if worker.time[x] == '08:00-15:45':
                times_list_with_count[9]+=1
            if worker.time[x] == '07:30-16:00':
                times_list_with_count[11]+=1
            if worker.time[x] == '06:00-14:00':
                times_list_with_count[13]+=1
            if worker.time[x] == '22:00-06:00':
                times_list_with_count[15]+=1
            if worker.time[x] == '14:00-22:00':
                times_list_with_count[17]+=1
            if worker.time[x] == '14:30-23:00':
                times_list_with_count[19]+=1
            if worker.time[x] == '10:00-18:30':
                times_list_with_count[21]+=1
            if worker.time[x] == '06:00-13:30':
                times_list_with_count[23]+=1
            if worker.time[x] == '06:00-14:30':
                times_list_with_count[25]+=1
            if worker.time[x] == '13:00-21:30':
                times_list_with_count[27]+=1
            if worker.time[x] == '08:00-16:00':
                times_list_with_count[29]+=1
            if worker.time[x] == '14:00-21:00':
                times_list_with_count[31]+=1
            if worker.time[x] == '07:00-15:00':
                times_list_with_count[33]+=1
            if worker.time[x] == '07:30-15:00':
                times_list_with_count[35]+=1
            if worker.time[x] == '10:00-17:30':
                times_list_with_count[37]+=1
            if worker.time[x] == '13:00-20:30':
                times_list_with_count[39]+=1
            if worker.time[x] == '14:30-22:00':
                times_list_with_count[41]+=1
            x+=1
        x=0


        # Sukurti reikiamus šablonus į work soberį, yra šablonų, kurie neatitinka vikarinos duomenų. Ant pavadinimų dėti laikus. MK 2 pam. 6.00-14.30, 14.30-23.00
        if times_list_with_count[29] == 1 and 18 <= times_list_with_count[7] <= 24:
            worker.template = '8.00-17.00, 08.00-16.00'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '8.00-17.00, 08.00-16.00')
            # print(template_object_list_temp[0].name, template_object_list_temp[0].time, template_object_list_temp[1].name, template_object_list_temp[1].time," nauji templates")
            # sys.exit()

        if 18 <= times_list_with_count[11] <= 24: #
            worker.template = 'Įrengimų priežiūra dieninė, 7.30-16.00' 
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Įrengimų priežiūra dieninė, 7.30-16.00')

        if 18 <= times_list_with_count[1] <= 24:
            worker.template = '7.00-16.00'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '7.00-16.00')

        if 18 <= times_list_with_count[3] <= 24:
            worker.template = '8.00 -12.00'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '8.00 -12.00')

        if 18 <= times_list_with_count[7] <= 24:
            worker.template = '8.00-17.00'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '8.00-17.00')
        
        if 16 <= times_list_with_count[7] <= 20 and 2 <= times_list_with_count[9] <= 5:
            worker.template = 'Administracija'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Administracija')

        if 10 <= times_list_with_count[25] <= 14 and 8 <= times_list_with_count[19] <= 12: 
            worker.template = 'MK 2 pam.'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'MK 2 pam.')

        if 8 <= times_list_with_count[15] <= 12 and 6 <= times_list_with_count[13] <= 10 and 3 <= times_list_with_count[17] <= 7: 
            worker.template = 'Naktinė'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Naktinė')

        if 7 <= times_list_with_count[17] <= 11 and 8 <= times_list_with_count[15] <= 12 and 3 <= times_list_with_count[11] <= 7: 
            worker.template = 'Popietinė'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Popietinė')

        if 14 <= times_list_with_count[19] <= 16 and 8 <= times_list_with_count[25] <= 12: 
            worker.template = 'Popietinė 2 pam.'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Popietinė 2 pam.')

        if 8 <= times_list_with_count[21] <= 10 and 5 <= times_list_with_count[27] <= 9 and 3 <= times_list_with_count[11] <= 6: 
            worker.template = 'Popietinė (Logistika)'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Popietinė (Logistika)')

        if 8 <= times_list_with_count[13] <= 11 and 6 <= times_list_with_count[17] <= 10 and 4 <= times_list_with_count[15] <= 8: 
            worker.template = 'Rytinė'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Rytinė')

        if 10 <= times_list_with_count[27] <= 14 and 8 <= times_list_with_count[19] <= 12:
            worker.template = 'Rytinė 2 pam.'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Rytinė 2 pam.')

        if 7 <= times_list_with_count[11] <= 11 and 6 <= times_list_with_count[19] <= 10 and 3 <= times_list_with_count[23] <= 7: 
            worker.template = 'Rytinė (Logistika)'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Rytinė (Logistika)')

        if 7 <= times_list_with_count[27] <= 10 and 6 <= times_list_with_count[11] <= 10 and 4 <= times_list_with_count[21] <= 6: 
            worker.template = 'Vakarinė (Logistika)'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Vakarinė (Logistika)')

        if 3 <= times_list_with_count[13] <= 7 and 7 <= times_list_with_count[15] <= 11 and 7 <= times_list_with_count[17] <= 11:
            worker.template = '14.00-22.00, 22.00-06.00, 6.00-14.00'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '14.00-22.00, 22.00-06.00, 6.00-14.00')

        if 8 <= times_list_with_count[13] <= 12 and 10 <= times_list_with_count[17] <= 14:
            worker.template = '6.00-14.00, 14.00-22.00'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '6.00-14.00, 14.00-22.00')

        if 7 <= times_list_with_count[11] <= 11 and 6 <= times_list_with_count[21] <= 10 and 3 <= times_list_with_count[27] <= 7: 
            worker.template = '7.30-16.00, 10.00-18.30, 13.00-21.30'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '7.30-16.00, 10.00-18.30, 13.00-21.30')                                                                             ###  ### 
                                                                                                                                                                                                
        if 8 <= times_list_with_count[11] <= 10 and 3 <= times_list_with_count[21] <= 7 and 7 <= times_list_with_count[23] <= 11:
            worker.template = '13.00-21.30, 7.30-16.00, 10.00-18.30'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '13.00-21.30, 7.30-16.00, 10.00-18.30')                                                                             ##    ##

        if 5 <= times_list_with_count[15] <= 6 and 8 <= times_list_with_count[13] <= 9 and 3 <= times_list_with_count[17] <= 4:
            worker.template = '22.00-6.00, 6.00-14.00, 14.00-22.00 ver. 2'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '22.00-6.00, 6.00-14.00, 14.00-22.00 ver. 2')

        if 5 <= times_list_with_count[15] <= 6 and 8 <= times_list_with_count[13] <= 9 and 3 <= times_list_with_count[17] <= 4:
            worker.template = '22.00-6.00, 6.00-14.00, 14.00-22.00 ver. 2'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '22.00-6.00, 6.00-14.00, 14.00-22.00 ver. 2')

        if 8 <= times_list_with_count[25] <= 10 and 9 <= times_list_with_count[19] <= 11 and 1 <= times_list_with_count[23] <= 2:
            worker.template = '6.00-14.30, 14.30-23.00, 6.00-13.30'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '6.00-14.30, 14.30-23.00, 6.00-13.30')

        if 6 <= times_list_with_count[13] <= 8 and 8 <= times_list_with_count[17] <= 10 and 4 <= times_list_with_count[15] <= 6:
            worker.template = 'Naktinė 6.00-14.00, 14.00-22.00, 22.00-06.00'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Naktinė 6.00-14.00, 14.00-22.00, 22.00-06.00')

        if 9 <= times_list_with_count[19] <= 11 and 9 <= times_list_with_count[25] <= 11 and 1 <= times_list_with_count[41] <= 2:
            worker.template = 'Rytinė 2 pam. 14.30-23.00, 6.00-14.30, 14.30-22.00'
            if (worker.time[0] == '14:30-23:00' or worker.time[0] == '') and (worker.time[1] == '14:30-23:00' or worker.time[2] == ''):
                template_object_list_temp_par = template_object_list_temp
                template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Rytinė 2 pam. 14.30-23.00, 6.00-14.30, 14.30-22.00')
            if (worker.time[0] == '06:00-14:30' or worker.time[0] == '') and (worker.time[1] == '06:00-14:30' or worker.time[2] == ''):
                template_object_list_temp_par = template_object_list_temp
                template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Rytinė 2 pam. 6.00-14.30, 14.30-23.00, 14.30-22.00')

        if 9 <= times_list_with_count[19] <= 11 and 9 <= times_list_with_count[25] <= 11:
            worker.template = 'Rytinė 2 pam. 14.30-23.00, 6.00-14.30'
            if (worker.time[0] == '14:30-23:00' or worker.time[0] == '') and (worker.time[1] == '14:30-23:00' or worker.time[2] == ''):
                template_object_list_temp_par = template_object_list_temp
                template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Rytinė 2 pam. 14.30-23.00, 6.00-14.30')
            if (worker.time[0] == '06:00-14:30' or worker.time[0] == '') and (worker.time[1] == '06:00-14:30' or worker.time[2] == ''):
                template_object_list_temp_par = template_object_list_temp
                template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, 'Rytinė 2 pam. 6.00-14.30, 14.30-23.00')

        if 8 <= times_list_with_count[19] <= 10 and 9 <= times_list_with_count[13] <= 11 and 1 <= times_list_with_count[31] <= 2:
            worker.template = '14.00-22.00, 6.00-14.00, 14.00-21.00'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '14.00-22.00, 6.00-14.00, 14.00-21.00')

        if 8 <= times_list_with_count[21] <= 10 and 9 <= times_list_with_count[27] <= 11 and 3 <= times_list_with_count[2] <= 5 and 1 <= times_list_with_count[33] <= 2:
            worker.template = '10.00-18.30, 13.00-21.30, 07.00-16.00, 07.00-15.00'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '10.00-18.30, 13.00-21.30, 07.00-16.00, 07.00-15.00')

        if 8 <= times_list_with_count[17] <= 10 and 9 <= times_list_with_count[13] <= 10:
            worker.template = '14.00-22.00, 6.00-14.00'
            if (worker.time[0] == '14:00-22:00' or worker.time[0] == '') and (worker.time[1] == '14:00-22:00' or worker.time[2] == ''):
                template_object_list_temp_par = template_object_list_temp
                template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '14.00-22.00, 6.00-14.00')
            if (worker.time[0] == '06:00-14:00' or worker.time[0] == '') and (worker.time[1] == '06:00-14:00' or worker.time[2] == ''):
                template_object_list_temp_par = template_object_list_temp
                template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '6.00-14.00, 14.00-22.00')


        if 7 <= times_list_with_count[21] <= 8 and 8 <= times_list_with_count[27] <= 9 and 3 <= times_list_with_count[11] <= 4 and 1 <= times_list_with_count[11] <= 2:
            worker.template = '10.00-18.30, 13.00-21.30, 7.30-16.00, 7.30-15.00'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '10.00-18.30, 13.00-21.30, 7.30-16.00, 7.30-15.00')

        if 7 <= times_list_with_count[21] <= 8 and 8 <= times_list_with_count[27] <= 9 and 3 <= times_list_with_count[11] <= 4 and 1 <= times_list_with_count[35] <= 2:
            worker.template = '10.00-18.30, 13.00-21.30, 07.30-16.00, 07.30-15.00'
            template_object_list_temp_par = template_object_list_temp
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '10.00-18.30, 13.00-21.30, 07.30-16.00, 07.30-15.00')

        if 7 <= times_list_with_count[19] <= 8 and 8 <= times_list_with_count[13] <= 9 and 4 <= times_list_with_count[21] <= 5 and 1 <= times_list_with_count[37] <= 2:
            worker.template = '14.30-23.00, 06.00-14.00, 10.00-18.30, 10.00-17.30'
            template_object_list_temp_par = template_object_list_temp 
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '14.30-23.00, 06.00-14.00, 10.00-18.30, 10.00-17.30')

        if 7 <= times_list_with_count[11] <= 8 and 8 <= times_list_with_count[21] <= 9 and 3 <= times_list_with_count[27] <= 4 and 1 <= times_list_with_count[39] <= 2:
            worker.template = '07.30-16.00, 10.00-18.30, 13.00-21.30, 13.00-20.30'
            template_object_list_temp_par = template_object_list_temp 
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '07.30-16.00, 10.00-18.30, 13.00-21.30, 13.00-20.30')

        if 7 <= times_list_with_count[27] <= 8 and 8 <= times_list_with_count[11] <= 9 and 3 <= times_list_with_count[21] <= 4 and 1 <= times_list_with_count[37] <= 2:
            worker.template = '13.00-21.30, 07.30-16.00, 10.00-18.30, 10.00-17.30'
            template_object_list_temp_par = template_object_list_temp 
            template_object_list_temp = create_newer_templates(worker, template_object_list_temp_par, '13.00-21.30, 07.30-16.00, 10.00-18.30, 10.00-17.30')

        

        times_list_with_count = ['07:00-16:00', 0, #1
                                 '08:00-12:00', 0, #3
                                 '08:00-16:30', 0, #5
                                 '08:00-17:00', 0, #7
                                 '08:00-15:45', 0, #9
                                 '07:30-16:00', 0, #11
                                 '06:00-14:00', 0, #13
                                 '22:00-06:00', 0, #15
                                 '14:00-22:00', 0, #17
                                 '14:30-23:00', 0, #19
                                 '10:00-18:30', 0, #21
                                 '06:00-13:30', 0, #23
                                 '06:00-14:30', 0, #25
                                 '13:00-21:30', 0, #27
                                 '08:00-16:00', 0, #29
                                 '14:00-21:00', 0, #31
                                 '07:00-15:00', 0, #33
                                 '07:30-15:00', 0, #35
                                 '10:00-17:30', 0, #37
                                 '13:00-20:30', 0, #39
                                 '14:30-22:00', 0] #41

    # for template in template_object_list_temp:
    #     print(template.name, template.time)
    # template_object_list_temp_par = template_object_list_temp
    # template_object_list_temp = fix_two_repeating_templates(template_object_list_temp_par)

    i = 0
    while i < len(template_object_list_temp):
        if i % 2 == 0:
            if i + 1 <= len(template_object_list_temp):
                template = Template(name = template_object_list_temp[i].name, time = get_time2())

                template_object_list_temp.insert(i + 1, template)

        i = i + 1


    return template_object_list_temp


def create_newer_templates(worker, template_object_list_temp_par, name):
    # patiktinti ar jau yra toks šablonas 
    temp_list = template_object_list_temp_par
    templatae = []

    template = Template(name = name, time = worker.time)
    # print(worker.time)
    if template not in template_object_list_temp_par:
        temp_list.append(template)


    return temp_list


def first_friday_of_month(year: int, month: int) -> datetime.date:
    """Returns the date of the first Friday of the given year and month."""
    first_day = datetime.date(year, month, 1)
    days_to_friday = (4 - first_day.weekday()) % 7
    return first_day + datetime.timedelta(days=days_to_friday)


# Patikrinama ar duoti metai yra keliamieji metai
def is_leap_year(year):

    year = int(year)

    if year % 4 == 0:
        if year % 100 == 0:
            if year % 400 == 0:
                return True
            else:
                return False
        else:
            return True
    else:
        return False


# Randama mėnesio ir dienų skaičius
def get_current_year_and_month():
    now = datetime.datetime.now()
    year = now.year
    month = now.month
    return year, month


# Randamas darbo laikas pagal tabelio numerį
def find_worker_time(sorted_unique_workers, tabel_nr):
    time = []
    for worker in sorted_unique_workers:
        if worker.tabel_number == tabel_nr:
            time = worker.time

    return time


# Randama darbuotojo pavardė
def find_worker_surname(sorted_unique_workers, tabel_nr):
    surname = ""
    for worker in sorted_unique_workers:
        if worker.tabel_number == tabel_nr:
            surname = worker.surname

    return surname


# Randamas dienų skaičius mėnesyje
def days_in_a_month():
    from calendar import monthrange

    now = datetime.datetime.now()
    year = now.year
    month = now.month

    return monthrange(year, month)[1]


# Randami grupių pavadinimai
def find_groups(textdata8):
    groups = []
    for i, line in enumerate(textdata8):
        if line != "":
            if line[0].isalpha():
                if  i + 4 < len(textdata8):
                    if textdata8[i+4] != "" and textdata8[i-1] != "":
                        if textdata8[i+4][0].isalpha() and textdata8[i-1].isdigit():
                            groups.append(line)

    return groups


# Randamas darbuotojo tabelio numeris pagal vardą
def find_tabel_number_by_name(sorted_unique_workers, name_surname):
    for worker in sorted_unique_workers:
        if worker.name == name_surname.split('_')[0] and worker.surname == name_surname.split('_')[1]:
            tabel_number = worker.tabel_number

    return tabel_number


# Pataisomi grupių pavadinimai
def find_current_group(current_group):
    group = "empty"
    if current_group == 'Administracija_Finansai':
        group = 'Finansai - Administracija'
    elif current_group == 'Administracija_Gamyba':
        group = 'Gamyba - Administracija'
    elif current_group == 'Administracija_Įrangos gamyba':
        group = 'Įrangos gamyba - Administracija'
    elif current_group == 'Administracija_Kokybė':
        group = 'Kokybė - Administracija'
    elif current_group == 'Administracija_Logistika':
        group = 'Logistika - Administracija'
    elif current_group == 'Administracija_Nuolatinis tob. ir sauga':
        group = 'NT ir saugos - Administracija'
    elif current_group == 'Administracija_Personalas':
        group = 'Personalas - Administracija'
    elif current_group == 'Administracija_Pirkimai':
        group = 'Pirkimai - Administracija'
    elif current_group == 'Administracija_Techninis':
        group = 'Techninis - Administracija'
    elif current_group == 'Administracija_Vadovai':
        group = 'Vadovai'
    elif current_group == 'Didžioji kalvė':
        group = 'Alkūniniai velenai'
    elif current_group == 'Įrengimų priežiūra':
        group = 'Įrengimų priežiūra'
    elif current_group == 'Logistika':
        group = 'Logistika - Administracija'
    elif current_group == 'Mažoji kalvė':
        group = 'Mažoji kalvė'
    elif current_group == 'Mechaninis apdirbimas':
        group = 'Mechaninis apdirbimas'
    elif current_group == 'Personalas_Pagalbiniai darbininkai':
        group = 'Kiti pagalbiniai darbininkai'
    elif current_group == 'Pirkimai_Sandėlio darbuotojai':
        group = 'Pirkimai - Administracija'
    elif current_group == 'Terminis apdirbimas':
        group = 'Terminis apdirbimas'
    elif current_group == 'Kokybė':
        group = 'Kokybė'
    elif current_group == 'Įrangos gamyba':
        group = 'Įrangos gamyba'

    return group


# Suformatuojamas data žodynas (dictionary), pagal dienų skaičių ir paduodamas writetoexcel funkcijai atspausdinimui
def formatdataframe(days, workers_list_full, template_object_list):
    data1 = {}

    if days == 28:
        data = {
                'Darbuotojo ID': [worker.tabel_number for worker in workers_list_full],
                'Vardas': [worker.name for worker in workers_list_full],
                'Pavardė': [worker.surname for worker in workers_list_full],
                'Padaliniai': [worker.division for worker in workers_list_full],
                'Šablonas': [worker.template for worker in workers_list_full],
                'Reikšmė': [worker.value for worker in workers_list_full],
                '1': [worker.time[0] for worker in workers_list_full],
                '2': [worker.time[1] for worker in workers_list_full],
                '3': [worker.time[2] for worker in workers_list_full],
                '4': [worker.time[3] for worker in workers_list_full],
                '5': [worker.time[4] for worker in workers_list_full],
                '6': [worker.time[5] for worker in workers_list_full],
                '7': [worker.time[6] for worker in workers_list_full],
                '8': [worker.time[7] for worker in workers_list_full],
                '9': [worker.time[8] for worker in workers_list_full],
                '10': [worker.time[9] for worker in workers_list_full],
                '11': [worker.time[10] for worker in workers_list_full],
                '12': [worker.time[11] for worker in workers_list_full],
                '13': [worker.time[12] for worker in workers_list_full],
                '14': [worker.time[13] for worker in workers_list_full],
                '15': [worker.time[14] for worker in workers_list_full],
                '16': [worker.time[15] for worker in workers_list_full],
                '18': [worker.time[16] for worker in workers_list_full],
                '17': [worker.time[17] for worker in workers_list_full],
                '19': [worker.time[18] for worker in workers_list_full],
                '20': [worker.time[19] for worker in workers_list_full],
                '21': [worker.time[20] for worker in workers_list_full],
                '22': [worker.time[21] for worker in workers_list_full],
                '23': [worker.time[22] for worker in workers_list_full],
                '24': [worker.time[23] for worker in workers_list_full],
                '25': [worker.time[24] for worker in workers_list_full],
                '26': [worker.time[25] for worker in workers_list_full],
                '27': [worker.time[26] for worker in workers_list_full],
                '28': [worker.time[27] for worker in workers_list_full],
                }
        data_templates = {
            'Šablonas': [template.name for template in template_object_list],
            '1': [template.time[0] for template in template_object_list],
            '2': [template.time[1] for template in template_object_list],
            '3': [template.time[2] for template in template_object_list],
            '4': [template.time[3] for template in template_object_list],
            '5': [template.time[4] for template in template_object_list],
            '6': [template.time[5] for template in template_object_list],
            '7': [template.time[6] for template in template_object_list],
            '8': [template.time[7] for template in template_object_list],
            '9': [template.time[8] for template in template_object_list],
            '10': [template.time[9] for template in template_object_list],
            '11': [template.time[10] for template in template_object_list],
            '12': [template.time[11] for template in template_object_list],
            '13': [template.time[12] for template in template_object_list],
            '14': [template.time[13] for template in template_object_list],
            '15': [template.time[14] for template in template_object_list],
            '16': [template.time[15] for template in template_object_list],
            '17': [template.time[16] for template in template_object_list],
            '18': [template.time[17] for template in template_object_list],
            '19': [template.time[18] for template in template_object_list],
            '20': [template.time[19] for template in template_object_list],
            '21': [template.time[20] for template in template_object_list],
            '22': [template.time[21] for template in template_object_list],
            '23': [template.time[22] for template in template_object_list],
            '24': [template.time[23] for template in template_object_list],
            '25': [template.time[24] for template in template_object_list],
            '26': [template.time[25] for template in template_object_list],
            '27': [template.time[26] for template in template_object_list],
            '28': [template.time[27] for template in template_object_list]
            }

    if days == 29:
        data = {
                'Darbuotojo ID': [worker.tabel_number for worker in workers_list_full],
                'Vardas': [worker.name for worker in workers_list_full],
                'Pavardė': [worker.surname for worker in workers_list_full],
                'Padaliniai': [worker.division for worker in workers_list_full],
                'Šablonas': [worker.template for worker in workers_list_full],
                'Reikšmė': [worker.value for worker in workers_list_full],
                '1': [worker.time[0] for worker in workers_list_full],
                '2': [worker.time[1] for worker in workers_list_full],
                '3': [worker.time[2] for worker in workers_list_full],
                '4': [worker.time[3] for worker in workers_list_full],
                '5': [worker.time[4] for worker in workers_list_full],
                '6': [worker.time[5] for worker in workers_list_full],
                '7': [worker.time[6] for worker in workers_list_full],
                '8': [worker.time[7] for worker in workers_list_full],
                '9': [worker.time[8] for worker in workers_list_full],
                '10': [worker.time[9] for worker in workers_list_full],
                '11': [worker.time[10] for worker in workers_list_full],
                '12': [worker.time[11] for worker in workers_list_full],
                '13': [worker.time[12] for worker in workers_list_full],
                '14': [worker.time[13] for worker in workers_list_full],
                '15': [worker.time[14] for worker in workers_list_full],
                '16': [worker.time[15] for worker in workers_list_full],
                '18': [worker.time[16] for worker in workers_list_full],
                '17': [worker.time[17] for worker in workers_list_full],
                '19': [worker.time[18] for worker in workers_list_full],
                '20': [worker.time[19] for worker in workers_list_full],
                '21': [worker.time[20] for worker in workers_list_full],
                '22': [worker.time[21] for worker in workers_list_full],
                '23': [worker.time[22] for worker in workers_list_full],
                '24': [worker.time[23] for worker in workers_list_full],
                '25': [worker.time[24] for worker in workers_list_full],
                '26': [worker.time[25] for worker in workers_list_full],
                '27': [worker.time[26] for worker in workers_list_full],
                '28': [worker.time[27] for worker in workers_list_full],
                '29': [worker.time[28] for worker in workers_list_full],
                }

        data_templates = {
            'Šablonas': [template.name for template in template_object_list],
            '1': [template.time[0] for template in template_object_list],
            '2': [template.time[1] for template in template_object_list],
            '3': [template.time[2] for template in template_object_list],
            '4': [template.time[3] for template in template_object_list],
            '5': [template.time[4] for template in template_object_list],
            '6': [template.time[5] for template in template_object_list],
            '7': [template.time[6] for template in template_object_list],
            '8': [template.time[7] for template in template_object_list],
            '9': [template.time[8] for template in template_object_list],
            '10': [template.time[9] for template in template_object_list],
            '11': [template.time[10] for template in template_object_list],
            '12': [template.time[11] for template in template_object_list],
            '13': [template.time[12] for template in template_object_list],
            '14': [template.time[13] for template in template_object_list],
            '15': [template.time[14] for template in template_object_list],
            '16': [template.time[15] for template in template_object_list],
            '17': [template.time[16] for template in template_object_list],
            '18': [template.time[17] for template in template_object_list],
            '19': [template.time[18] for template in template_object_list],
            '20': [template.time[19] for template in template_object_list],
            '21': [template.time[20] for template in template_object_list],
            '22': [template.time[21] for template in template_object_list],
            '23': [template.time[22] for template in template_object_list],
            '24': [template.time[23] for template in template_object_list],
            '25': [template.time[24] for template in template_object_list],
            '26': [template.time[25] for template in template_object_list],
            '27': [template.time[26] for template in template_object_list],
            '28': [template.time[27] for template in template_object_list],
            '29': [template.time[28] for template in template_object_list]
            }
        
    if days == 30:
        data = {
                'Darbuotojo ID': [worker.tabel_number for worker in workers_list_full],
                'Vardas': [worker.name for worker in workers_list_full],
                'Pavardė': [worker.surname for worker in workers_list_full],
                'Padaliniai': [worker.division for worker in workers_list_full],
                'Šablonas': [worker.template for worker in workers_list_full],
                'Reikšmė': [worker.value for worker in workers_list_full],
                '1': [worker.time[0] for worker in workers_list_full],
                '2': [worker.time[1] for worker in workers_list_full],
                '3': [worker.time[2] for worker in workers_list_full],
                '4': [worker.time[3] for worker in workers_list_full],
                '5': [worker.time[4] for worker in workers_list_full],
                '6': [worker.time[5] for worker in workers_list_full],
                '7': [worker.time[6] for worker in workers_list_full],
                '8': [worker.time[7] for worker in workers_list_full],
                '9': [worker.time[8] for worker in workers_list_full],
                '10': [worker.time[9] for worker in workers_list_full],
                '11': [worker.time[10] for worker in workers_list_full],
                '12': [worker.time[11] for worker in workers_list_full],
                '13': [worker.time[12] for worker in workers_list_full],
                '14': [worker.time[13] for worker in workers_list_full],
                '15': [worker.time[14] for worker in workers_list_full],
                '16': [worker.time[15] for worker in workers_list_full],
                '18': [worker.time[16] for worker in workers_list_full],
                '17': [worker.time[17] for worker in workers_list_full],
                '19': [worker.time[18] for worker in workers_list_full],
                '20': [worker.time[19] for worker in workers_list_full],
                '21': [worker.time[20] for worker in workers_list_full],
                '22': [worker.time[21] for worker in workers_list_full],
                '23': [worker.time[22] for worker in workers_list_full],
                '24': [worker.time[23] for worker in workers_list_full],
                '25': [worker.time[24] for worker in workers_list_full],
                '26': [worker.time[25] for worker in workers_list_full],
                '27': [worker.time[26] for worker in workers_list_full],
                '28': [worker.time[27] for worker in workers_list_full],
                '29': [worker.time[28] for worker in workers_list_full],
                '30': [worker.time[29] for worker in workers_list_full]
                }
        
        data_templates = {
            'Šablonas': [template.name for template in template_object_list],
            '1': [template.time[0] for template in template_object_list],
            '2': [template.time[1] for template in template_object_list],
            '3': [template.time[2] for template in template_object_list],
            '4': [template.time[3] for template in template_object_list],
            '5': [template.time[4] for template in template_object_list],
            '6': [template.time[5] for template in template_object_list],
            '7': [template.time[6] for template in template_object_list],
            '8': [template.time[7] for template in template_object_list],
            '9': [template.time[8] for template in template_object_list],
            '10': [template.time[9] for template in template_object_list],
            '11': [template.time[10] for template in template_object_list],
            '12': [template.time[11] for template in template_object_list],
            '13': [template.time[12] for template in template_object_list],
            '14': [template.time[13] for template in template_object_list],
            '15': [template.time[14] for template in template_object_list],
            '16': [template.time[15] for template in template_object_list],
            '17': [template.time[16] for template in template_object_list],
            '18': [template.time[17] for template in template_object_list],
            '19': [template.time[18] for template in template_object_list],
            '20': [template.time[19] for template in template_object_list],
            '21': [template.time[20] for template in template_object_list],
            '22': [template.time[21] for template in template_object_list],
            '23': [template.time[22] for template in template_object_list],
            '24': [template.time[23] for template in template_object_list],
            '25': [template.time[24] for template in template_object_list],
            '26': [template.time[25] for template in template_object_list],
            '27': [template.time[26] for template in template_object_list],
            '28': [template.time[27] for template in template_object_list],
            '29': [template.time[28] for template in template_object_list],
            '30': [template.time[29] for template in template_object_list]
            }
    
    if days == 31:
        data = {
                'Darbuotojo ID': [worker.tabel_number for worker in workers_list_full],
                'Vardas': [worker.name for worker in workers_list_full],
                'Pavardė': [worker.surname for worker in workers_list_full],
                'Padaliniai': [worker.division for worker in workers_list_full],
                'Šablonas': [worker.template for worker in workers_list_full],
                'Reikšmė': [worker.value for worker in workers_list_full],
                '1': [worker.time[0] for worker in workers_list_full],
                '2': [worker.time[1] for worker in workers_list_full],
                '3': [worker.time[2] for worker in workers_list_full],
                '4': [worker.time[3] for worker in workers_list_full],
                '5': [worker.time[4] for worker in workers_list_full],
                '6': [worker.time[5] for worker in workers_list_full],
                '7': [worker.time[6] for worker in workers_list_full],
                '8': [worker.time[7] for worker in workers_list_full],
                '9': [worker.time[8] for worker in workers_list_full],
                '10': [worker.time[9] for worker in workers_list_full],
                '11': [worker.time[10] for worker in workers_list_full],
                '12': [worker.time[11] for worker in workers_list_full],
                '13': [worker.time[12] for worker in workers_list_full],
                '14': [worker.time[13] for worker in workers_list_full],
                '15': [worker.time[14] for worker in workers_list_full],
                '16': [worker.time[15] for worker in workers_list_full],
                '18': [worker.time[16] for worker in workers_list_full],
                '17': [worker.time[17] for worker in workers_list_full],
                '19': [worker.time[18] for worker in workers_list_full],
                '20': [worker.time[19] for worker in workers_list_full],
                '21': [worker.time[20] for worker in workers_list_full],
                '22': [worker.time[21] for worker in workers_list_full],
                '23': [worker.time[22] for worker in workers_list_full],
                '24': [worker.time[23] for worker in workers_list_full],
                '25': [worker.time[24] for worker in workers_list_full],
                '26': [worker.time[25] for worker in workers_list_full],
                '27': [worker.time[26] for worker in workers_list_full],
                '28': [worker.time[27] for worker in workers_list_full],
                '29': [worker.time[28] for worker in workers_list_full],
                '30': [worker.time[29] for worker in workers_list_full],
                '31': [worker.time[29] for worker in workers_list_full]
                }
        
        data_templates = {
            'Šablonas': [template.name for template in template_object_list],
            '1': [template.time[0] for template in template_object_list],
            '2': [template.time[1] for template in template_object_list],
            '3': [template.time[2] for template in template_object_list],
            '4': [template.time[3] for template in template_object_list],
            '5': [template.time[4] for template in template_object_list],
            '6': [template.time[5] for template in template_object_list],
            '7': [template.time[6] for template in template_object_list],
            '8': [template.time[7] for template in template_object_list],
            '9': [template.time[8] for template in template_object_list],
            '10': [template.time[9] for template in template_object_list],
            '11': [template.time[10] for template in template_object_list],
            '12': [template.time[11] for template in template_object_list],
            '13': [template.time[12] for template in template_object_list],
            '14': [template.time[13] for template in template_object_list],
            '15': [template.time[14] for template in template_object_list],
            '16': [template.time[15] for template in template_object_list],
            '17': [template.time[16] for template in template_object_list],
            '18': [template.time[17] for template in template_object_list],
            '19': [template.time[18] for template in template_object_list],
            '20': [template.time[19] for template in template_object_list],
            '21': [template.time[20] for template in template_object_list],
            '22': [template.time[21] for template in template_object_list],
            '23': [template.time[22] for template in template_object_list],
            '24': [template.time[23] for template in template_object_list],
            '25': [template.time[24] for template in template_object_list],
            '26': [template.time[25] for template in template_object_list],
            '27': [template.time[26] for template in template_object_list],
            '28': [template.time[27] for template in template_object_list],
            '29': [template.time[28] for template in template_object_list],
            '30': [template.time[29] for template in template_object_list],
            '31': [template.time[30] for template in template_object_list]
            }
        
    if days == 31:
        data1 = {
                'Darbuotojo ID': [workers_list_full[0].tabel_number, workers_list_full[1].tabel_number],
                'Vardas': [workers_list_full[0].name, workers_list_full[1].name],
                'Pavardė': [workers_list_full[0].surname, workers_list_full[1].surname],
                'Padaliniai': [workers_list_full[0].division, workers_list_full[1].division],
                'Šablonas': [workers_list_full[0].template, workers_list_full[1].template],
                'Reikšmė': [workers_list_full[0].value, workers_list_full[1].value],
                '1': [workers_list_full[0].time[0], workers_list_full[1].time[0]],
                '2': [workers_list_full[0].time[1], workers_list_full[1].time[1]],
                '3': [workers_list_full[0].time[2], workers_list_full[1].time[2]],
                '4': [workers_list_full[0].time[3], workers_list_full[1].time[3]],
                '5': [workers_list_full[0].time[4], workers_list_full[1].time[4]],
                '6': [workers_list_full[0].time[5], workers_list_full[1].time[5]],
                '7': [workers_list_full[0].time[6], workers_list_full[1].time[6]],
                '8': [workers_list_full[0].time[7], workers_list_full[1].time[7]],
                '9': [workers_list_full[0].time[8], workers_list_full[1].time[8]],
                '10': [workers_list_full[0].time[9], workers_list_full[1].time[9]],
                '11': [workers_list_full[0].time[10], workers_list_full[1].time[10]],
                '12': [workers_list_full[0].time[11], workers_list_full[1].time[11]],
                '13': [workers_list_full[0].time[12], workers_list_full[1].time[12]],
                '14': [workers_list_full[0].time[13], workers_list_full[1].time[13]],
                '15': [workers_list_full[0].time[14], workers_list_full[1].time[14]],
                '16': [workers_list_full[0].time[15], workers_list_full[1].time[15]],
                '18': [workers_list_full[0].time[16], workers_list_full[1].time[16]],
                '17': [workers_list_full[0].time[17], workers_list_full[1].time[17]],
                '19': [workers_list_full[0].time[18], workers_list_full[1].time[18]],
                '20': [workers_list_full[0].time[19], workers_list_full[1].time[19]],
                '21': [workers_list_full[0].time[20], workers_list_full[1].time[20]],
                '22': [workers_list_full[0].time[21], workers_list_full[1].time[21]],
                '23': [workers_list_full[0].time[22], workers_list_full[1].time[22]],
                '24': [workers_list_full[0].time[23], workers_list_full[1].time[23]],
                '25': [workers_list_full[0].time[24], workers_list_full[1].time[24]],
                '26': [workers_list_full[0].time[25], workers_list_full[1].time[25]],
                '27': [workers_list_full[0].time[26], workers_list_full[1].time[26]],
                '28': [workers_list_full[0].time[27], workers_list_full[1].time[27]],
                '29': [workers_list_full[0].time[28], workers_list_full[1].time[28]],
                '30': [workers_list_full[0].time[29], workers_list_full[1].time[29]],
                '31': [workers_list_full[0].time[30], workers_list_full[1].time[30]]
                }

    
    return data, data1, data_templates


# Sukuriama darbuotojo klasė patogiam darbuotojų duomenų įvedinimu. Veikia kaip šablonas
class Template:
    def __init__(self, name, time):
        self.name = name
        self.time = time

    def __eq__(self, other):
        if not isinstance(other, Template):
            return False

        return self.name == other.name and self.time == other.time

    def __hash__(self):
        return hash((self.name, self.time))


# Sukuriama papildyda darbuotojo klasė patogiam darbuotojų duomenų įvedinimu į excel 
class Worker_Modified:
    def __init__(self, tabel_number, name, surname, time_start, time_end, time, division, template, value):
        self.tabel_number = tabel_number
        self.name = name
        self.surname = surname
        self.time_start = time_start
        self.time_end = time_end
        self.time = time
        self.division = division
        self.template = template
        self.value = value

class Worker:
    def __init__(self, tabel_number, name, surname, time_start, time_end, time, division):
        self.tabel_number = tabel_number
        self.name = name
        self.surname = surname
        self.time_start = time_start
        self.time_end = time_end
        self.time = time
        self.division = division

# Pradedama main() funkcija, prieš tai python nuskaito klases, funkcijas ir kintamuosius
if __name__ == "__main__":
    initial_x = 0
    initial_y = 0
    main()